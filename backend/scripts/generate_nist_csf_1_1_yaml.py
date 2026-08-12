"""One-off generator for framework_mapping/nist_csf_1_1.yaml.

Not part of the application (see scripts/README.md). Mirrors
generate_nist_csf_yaml.py's pattern for the SAME framework at an EARLIER
version, which is the entire point of this file: ADR-0053 built
multi-version registry support and disclosed that no real framework in
this project had more than one version, so the mechanism had only ever
been exercised against a synthetic two-version fixture. NIST CSF 1.1 is
that real second version (ADR-0055).

Why CSF 1.1 and not an older version of some other framework: it is the
only candidate with no copyright constraint. NIST publications are US
Government works in the public domain, so the full Subcategory text can
be transcribed verbatim -- unlike ISO 27001, SOC 2, and PCI DSS, whose
earlier versions are as copyrighted as their current ones and would have
forced another titles-only/statement-only compromise (ADR-0024/0026/0027).

Sources, both fetched directly and verified rather than assumed:

1. Categories and Subcategories -- the official CSF 1.1 Core spreadsheet,
   https://www.nist.gov/document/2018-04-16frameworkv11core1xlsx
   Parsed here rather than hand-transcribed: it is the authoritative
   machine-readable Core, and parsing 108 subcategories is both more
   accurate and more reproducible than retyping them.
2. Function purpose statements -- "Framework for Improving Critical
   Infrastructure Cybersecurity, Version 1.1", NIST, April 16 2018,
   https://nvlpubs.nist.gov/nistpubs/CSWP/NIST.CSWP.04162018.pdf
   The Core spreadsheet does not carry Function definitions at all, so
   these five are transcribed verbatim (first sentence of each Function's
   definition in Section 2.1) from the PDF, confirmed genuine by checking
   the document's own title and "Version 1.1" markers before reading it.

Structure maps to this project's shared schema exactly as CSF 2.0 does:
Function -> Domain, Category -> Objective, Subcategory -> Practice. Like
2.0, CSF 1.1 defines no native maturity levels, so practices carry no
`mil` and the framework is scored by coverage.

Counts are asserted, not trusted: 5 Functions, 23 Categories, 108
Subcategories, matching NIST's own published totals for CSF 1.1. The
generator fails loudly if the parsed source disagrees.

Usage:
    python scripts/generate_nist_csf_1_1_yaml.py [path-to-core.xlsx]
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_PATH = REPO_ROOT / "framework_mapping" / "nist_csf_1_1.yaml"

CORE_XLSX_URL = "https://www.nist.gov/document/2018-04-16frameworkv11core1xlsx"
SOURCE_PDF_URL = "https://nvlpubs.nist.gov/nistpubs/CSWP/NIST.CSWP.04162018.pdf"

EXPECTED_FUNCTIONS = 5
EXPECTED_CATEGORIES = 23
EXPECTED_SUBCATEGORIES = 108

# Verbatim first sentence of each Function's definition, Section 2.1 of
# the CSF 1.1 PDF. CSF 1.1 has five Functions -- Govern (GV) does not
# exist until 2.0, which is the single most significant structural
# difference between the two versions and the reason a 1.1-pinned
# assessment cannot simply be re-scored against 2.0.
FUNCTION_PURPOSES = {
    "ID": (
        "Identify",
        "Develop an organizational understanding to manage cybersecurity risk to systems, "
        "people, assets, data, and capabilities.",
    ),
    "PR": (
        "Protect",
        "Develop and implement appropriate safeguards to ensure delivery of critical services.",
    ),
    "DE": (
        "Detect",
        "Develop and implement appropriate activities to identify the occurrence of a "
        "cybersecurity event.",
    ),
    "RS": (
        "Respond",
        "Develop and implement appropriate activities to take action regarding a detected "
        "cybersecurity incident.",
    ),
    "RC": (
        "Recover",
        "Develop and implement appropriate activities to maintain plans for resilience and to "
        "restore any capabilities or services that were impaired due to a cybersecurity "
        "incident.",
    ),
}

SCORING_NOTE = (
    "NIST CSF 1.1 does not natively define maturity levels for its Functions, Categories, or "
    "Subcategories (the Framework Implementation Tiers are an organization-wide risk-management "
    "characterization, explicitly NOT a maturity model -- see the CSF 1.1 source document, "
    "Section 2.2). Scores for this framework are a project-defined coverage measure: the "
    "fraction of a Function's Subcategories with accepted or edited evidence, computed by "
    "services/scoring_service.py's compute_domain_coverage. This is NOT part of the NIST "
    "standard itself and must always be presented as such."
)


def _clean(value: object) -> str:
    """Collapse the non-breaking spaces and hard line breaks the source
    spreadsheet's wrapped cells are full of."""
    return " ".join(str(value).replace("\xa0", " ").split())


def _split_once(text: str) -> tuple[str, str]:
    head, _, tail = text.partition(":")
    return head.strip(), tail.strip()


def parse_core(xlsx_path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    # Function/Category cells are populated only on the first row of each
    # group (the source uses vertically merged cells), so both are
    # forward-filled as the rows are walked.
    function_code: str | None = None
    category_key: tuple[str, str] | None = None
    domains: dict[str, dict] = {}

    for function_cell, category_cell, subcategory_cell, _references in sheet.iter_rows(
        min_row=2, values_only=True
    ):
        if function_cell:
            # e.g. "IDENTIFY (ID)"
            function_code = _clean(function_cell).rsplit("(", 1)[-1].rstrip(")").strip()
        if category_cell:
            title, purpose = _split_once(_clean(category_cell))
            category_key = (title, purpose)
        if not subcategory_cell:
            continue

        assert function_code is not None and category_key is not None
        practice_id, text = _split_once(_clean(subcategory_cell))

        domain = domains.setdefault(
            function_code, {"short_code": function_code, "objectives": {}}
        )
        objective = domain["objectives"].setdefault(
            category_key, {"title": category_key[0], "purpose": category_key[1], "practices": []}
        )
        objective["practices"].append({"id": practice_id, "text": text})

    workbook.close()

    ordered: list[dict] = []
    for code, domain in domains.items():
        full_name, purpose = FUNCTION_PURPOSES[code]
        ordered.append(
            {
                "short_code": code,
                "full_name": full_name,
                "purpose": purpose,
                "practices_populated": True,
                "objectives": [
                    {
                        "number": index,
                        "title": objective["title"],
                        "purpose": objective["purpose"],
                        "practices": objective["practices"],
                    }
                    for index, objective in enumerate(domain["objectives"].values(), start=1)
                ],
            }
        )
    return ordered


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/tmp/csf11core.xlsx")
    if not xlsx_path.exists():
        raise SystemExit(
            f"Core spreadsheet not found at {xlsx_path}. Download it first:\n"
            f"  curl -sL -A 'Mozilla/5.0' -o {xlsx_path} '{CORE_XLSX_URL}'"
        )

    domains = parse_core(xlsx_path)

    # Fail loudly on any disagreement with NIST's own published totals,
    # rather than emitting a plausible-looking file with silently missing
    # content -- the same discipline every other framework generator here
    # uses (see ADR-0018, ADR-0022, ADR-0029).
    category_count = sum(len(d["objectives"]) for d in domains)
    practice_count = sum(len(o["practices"]) for d in domains for o in d["objectives"])
    assert len(domains) == EXPECTED_FUNCTIONS, f"functions: {len(domains)}"
    assert category_count == EXPECTED_CATEGORIES, f"categories: {category_count}"
    assert practice_count == EXPECTED_SUBCATEGORIES, f"subcategories: {practice_count}"

    document = {
        "name": "NIST CSF 1.1",
        "full_name": "NIST Cybersecurity Framework",
        "version": "1.1",
        "source_title": (
            "Framework for Improving Critical Infrastructure Cybersecurity, Version 1.1"
        ),
        "source_publisher": "National Institute of Standards and Technology",
        "source_date": "2018-04-16",
        "source_url": SOURCE_PDF_URL,
        "retrieved_date": "2026-08-12",
        "total_practices_in_source": EXPECTED_SUBCATEGORIES,
        "scoring_model": "coverage",
        "mil_levels": [],
        "scoring_note": SCORING_NOTE,
        "domains": domains,
    }

    header = (
        "# GENERATED FILE. Do not hand-edit — regenerate via\n"
        "# backend/scripts/generate_nist_csf_1_1_yaml.py, which is the source of\n"
        "# truth for this file's content and carries the source citation.\n"
    )
    with OUTPUT_PATH.open("w", encoding="utf-8") as handle:
        handle.write(header)
        yaml.safe_dump(document, handle, sort_keys=False, allow_unicode=True, width=100)

    print(
        f"wrote {OUTPUT_PATH.relative_to(REPO_ROOT)}: "
        f"{len(domains)} functions, {category_count} categories, {practice_count} subcategories"
    )


if __name__ == "__main__":
    main()
