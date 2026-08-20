"""PDF/XLSX rendering of an already-computed DashboardReport (Sprint 7).

Sprint 6 (report_service.py) computed a DashboardReport: situation,
overall summary, MECE complication groups, and a prioritized resolution
list. This module is a pure rendering step on top of that already-real,
already-verified data — no new computation, no LLM narrative, and
nothing persisted server-side. See ADR-0013 for why export is scoped
this narrowly and why PDF and XLSX deliberately render different views
of the same data rather than the same layout twice.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Font as XlsxFont
from openpyxl.worksheet.worksheet import Worksheet

from compliance_platform.models.assessment import PracticeFindingStatus
from compliance_platform.models.report import DashboardReport, EvidenceCitation
from compliance_platform.models.schemas import TextProvenance
from compliance_platform.services.report_currency import (
    CURRENT_PAYLOAD_VERSION,
    report_digest,
)

# fpdf2's core fonts (Helvetica/Times/Courier) only reliably encode
# Latin-1. This project's framework source text is transcribed verbatim
# from DOE/NIST PDFs (c2m2-expert/nist-csf-expert skills) and contains
# occasional em dashes outside that range. Translate the common
# punctuation cases explicitly rather than let a rare character crash
# report generation or silently mangle into mojibake — the same
# explicit-failure-mode discipline the document-parsing skill requires
# of the ingestion parsers, applied here to the export path.
_PDF_CHAR_REPLACEMENTS = {
    "—": " - ",  # em dash
    "–": "-",  # en dash
    "‘": "'",
    "’": "'",
    "“": '"',
    "”": '"',
    "…": "...",
}


def _pdf_safe(text: str) -> str:
    for bad, good in _PDF_CHAR_REPLACEMENTS.items():
        text = text.replace(bad, good)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _mil_label(mil: int | None) -> str:
    return f"MIL{mil}" if mil is not None else "n/a"


def _status_label(status: PracticeFindingStatus) -> str:
    return status.value.replace("_", " ")


_PROVENANCE_LABELS = {
    TextProvenance.OCR: "OCR - approximate",
    TextProvenance.POSSIBLY_OCR: "may be OCR",
    TextProvenance.UNKNOWN: "provenance unrecorded",
}


def _provenance_tag(provenance: TextProvenance) -> str:
    """A tag only when there is something to say (ADR-0076).

    `exact` renders nothing, in the export exactly as on screen: a note
    on every ordinary citation is noise, and noise is what stops anyone
    reading the one that matters.
    """
    label = _PROVENANCE_LABELS.get(provenance)
    return f" [{label}]" if label else ""


def _evidence_citation_summary(citations: list[EvidenceCitation]) -> str:
    # IDs/status only, per EvidenceCitation's own docstring -- never the
    # underlying evidence text. [SUPERSEDED] (ADR-0050) surfaces in the
    # export itself, not just via a separate API call a reviewer would
    # have to think to make.
    def _one(c: EvidenceCitation) -> str:
        superseded_tag = " [SUPERSEDED]" if c.is_superseded else ""
        return (
            f"{c.document_id} ({c.review_status.value})"
            f"{superseded_tag}{_provenance_tag(c.text_provenance)}"
        )

    return "; ".join(_one(c) for c in citations)


# Segment colours for the review-progress bar, matching the dashboard's
# (ADR-0068). Rejected is deliberately grey rather than red: retrieval
# precision was measured at 0.012, so declining a proposal is the
# expected outcome for most of a queue, and a report that paints correct
# review work as an alarm misreports it -- on paper as much as on screen.
_REVIEW_SEGMENT_COLOURS = {
    "accepted": (16, 150, 105),
    "edited": (14, 132, 197),
    "rejected": (148, 158, 170),
    "pending": (222, 148, 22),
}

_BAR_WIDTH = 120.0
_BAR_HEIGHT = 3.2


def _score_label(score: float, scoring_model: str) -> str:
    """A domain score in the units of its own scoring model.

    Never a bare float. `domain_scores` is an ordinal MIL 0-3 under
    cumulative_mil and a 0.0-1.0 fraction under coverage (R-15), and a
    number printed without its unit is the same ambiguity ADR-0066
    refused to draw as a bar.
    """
    if scoring_model == "cumulative_mil":
        return f"MIL{score:.0f}"
    return f"{score * 100:.0f}% coverage"


def _score_column_header(scoring_model: str) -> str:
    if scoring_model == "cumulative_mil":
        return "Score (MIL 0-3, ordinal)"
    return "Score (coverage, 0.0-1.0)"


def _draw_progress_bar(pdf: FPDF, filled_fraction: float) -> None:
    """One horizontal bar at the current cursor, drawn as two rectangles.

    Rendered rather than described because the PDF is this project's
    narrative artifact (ADR-0013) and the tester's report was that the
    dashboard's numbers needed a shape. No charting dependency: a filled
    rectangle over a track is the whole requirement.
    """
    x = pdf.get_x()
    y = pdf.get_y()
    pdf.set_fill_color(226, 230, 235)
    pdf.rect(x, y, _BAR_WIDTH, _BAR_HEIGHT, style="F")
    if filled_fraction > 0:
        pdf.set_fill_color(79, 92, 216)
        pdf.rect(x, y, _BAR_WIDTH * min(filled_fraction, 1.0), _BAR_HEIGHT, style="F")
    pdf.ln(_BAR_HEIGHT + 2)


def _draw_review_bar(pdf: FPDF, segments: list[tuple[str, int]], total: int) -> None:
    """The review-status breakdown as one stacked bar (ADR-0068)."""
    if total <= 0:
        return
    x = pdf.get_x()
    y = pdf.get_y()
    pdf.set_fill_color(226, 230, 235)
    pdf.rect(x, y, _BAR_WIDTH, _BAR_HEIGHT, style="F")
    offset = 0.0
    for key, count in segments:
        if count <= 0:
            continue
        width = _BAR_WIDTH * (count / total)
        pdf.set_fill_color(*_REVIEW_SEGMENT_COLOURS[key])
        pdf.rect(x + offset, y, width, _BAR_HEIGHT, style="F")
        offset += width
    pdf.ln(_BAR_HEIGHT + 2)


def _line(pdf: FPDF, height: float, text: str) -> None:
    """Write one paragraph and reliably leave the cursor at the left
    margin on the next line. fpdf2's own default post-multi_cell cursor
    position is not guaranteed to be back at the left margin, and
    chaining calls without pinning it raised a real "not enough
    horizontal space to render a single character" error during testing
    — pinning new_x/new_y explicitly here is the fix, not a workaround
    for a one-off bug.
    """
    pdf.multi_cell(0, height, _pdf_safe(text), new_x=XPos.LMARGIN, new_y=YPos.NEXT)


def build_pdf_report(dashboard: DashboardReport) -> bytes:
    """The board-ready narrative artifact (US-6.2's Marcus persona):
    fixed prose in the same situation/complication/resolution order as
    the dashboard API, not a data dump. Every gap still shows its
    AI-proposed/pending flag so a reviewer can't mistake a proposed
    mapping for verified evidence in a document that has left the API.
    """
    pdf = FPDF(format="Letter")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    generated_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    s = dashboard.situation

    pdf.set_font("Helvetica", "B", 16)
    _line(pdf, 10, s.assessment_name)
    # Whose assessment this is, on the page itself (ADR-0063): an export
    # is the copy that leaves the database, and a compliance report that
    # does not name its subject organisation is one that can be filed
    # against the wrong client.
    if s.organization_name:
        _line(pdf, 6, s.organization_name)
    pdf.set_font("Helvetica", "", 10)
    _line(
        pdf,
        6,
        f"Framework: {s.framework_name}   |   Status: {s.status}   |   "
        f"Generated: {generated_at}",
    )
    # The finalization seal, printed so a copy of it leaves the database
    # with the report (models/report.py's Situation.finalization_seal).
    # A reader holding this PDF a year from now can check the value
    # against GET /assessments/{id}/verify and find out whether the
    # record still says what it said here.
    if s.finalization_seal:
        pdf.set_font("Courier", "", 7)
        _line(pdf, 5, f"Finalization seal (SHA-256): {s.finalization_seal}")
        pdf.set_font("Helvetica", "", 10)
    # The report digest (ADR-0077), printed on every export rather than
    # only on finalized ones. The seal above answers "has this immutable
    # record been altered?"; this answers "have the numbers moved since
    # this page was printed?", which is the question a draft export
    # raises and R-21 has named since Sprint 7.
    pdf.set_font("Courier", "", 7)
    _line(
        pdf,
        5,
        f"Report digest (SHA-256, v{CURRENT_PAYLOAD_VERSION}): {report_digest(dashboard)}",
    )
    pdf.set_font("Helvetica", "", 8)
    _line(
        pdf,
        4,
        "Check whether this report is still current: GET "
        f"/assessments/{s.assessment_id}/report-currency?digest=<the digest above>",
    )
    pdf.set_font("Helvetica", "", 10)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 13)
    _line(pdf, 8, "Situation")
    pdf.set_font("Helvetica", "", 10)
    _line(
        pdf,
        6,
        f"{s.total_evidence_links} evidence link(s) total: {s.accepted_count} accepted, "
        f"{s.edited_count} edited, {s.rejected_count} rejected, "
        f"{s.pending_ai_review_count} still pending human review "
        "(AI-proposed, not yet counted toward any score below).",
    )
    # The same breakdown as a shape (ADR-0068). The counts above are the
    # figures; this is how much of the assessment is still undecided,
    # which is what decides how far everything below can be trusted.
    _draw_review_bar(
        pdf,
        [
            ("accepted", s.accepted_count),
            ("edited", s.edited_count),
            ("rejected", s.rejected_count),
            ("pending", s.pending_ai_review_count),
        ],
        s.total_evidence_links,
    )
    if s.unpopulated_domains:
        _line(
            pdf,
            6,
            "Not yet transcribed into the platform's schema, excluded from scoring below: "
            + ", ".join(s.unpopulated_domains)
            + ".",
        )
    # What those counts MEAN (executive-reporting.mdc's "every number
    # needs a so what"). The counts above are the same figures the
    # dashboard shows; these are the same interpretation the dashboard
    # shows, computed once in report_service so the exported document and
    # the screen can never disagree about what the numbers imply.
    if s.so_what:
        pdf.ln(2)
        pdf.set_font("Helvetica", "I", 10)
        for sentence in s.so_what:
            _line(pdf, 5, f"- {sentence}")
        pdf.set_font("Helvetica", "", 10)
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 11)
    _line(pdf, 6, dashboard.overall.headline)
    pdf.ln(4)

    # Domain completion (ADR-0066), which until now existed only on
    # screen. Distinct from Complication below: that section lists
    # domains with at least one gap, so a fully-met domain never appears
    # in it, and a reader of the PDF alone could not tell a finished
    # domain from one nobody had started.
    if dashboard.domain_progress:
        pdf.set_font("Helvetica", "B", 13)
        _line(pdf, 8, "Domain Completion")
        pdf.set_font("Helvetica", "", 9)
        _line(
            pdf,
            5,
            "Bars show applicable practices met. Practices marked not applicable are excluded. "
            + (
                "They are not the maturity score: MIL is cumulative, so a domain reaches MIL2 "
                "only when every MIL1 practice is also met."
                if dashboard.overall.scoring_model == "cumulative_mil"
                else "This is the same measure as this framework's coverage score."
            ),
        )
        pdf.ln(1)
        for entry in dashboard.domain_progress:
            pdf.set_font("Helvetica", "B", 10)
            _line(
                pdf,
                5,
                f"{entry.short_code} - {entry.full_name}: {entry.met_practices}/"
                f"{entry.total_practices} practices, "
                f"{_score_label(entry.score, dashboard.overall.scoring_model)}",
            )
            _draw_progress_bar(pdf, entry.met_practices / entry.total_practices)
            # The reconciliation between a nearly-full bar and a low
            # score. Composed in report_service (ADR-0069) so this
            # document and the screen cannot word it differently.
            if entry.gate_note:
                pdf.set_font("Helvetica", "I", 9)
                _line(pdf, 5, f"  {entry.gate_note}")
            pdf.ln(1)
        pdf.ln(2)

    pdf.set_font("Helvetica", "B", 13)
    _line(pdf, 8, "Complication - Gaps by Domain")
    pdf.set_font("Helvetica", "", 10)
    if not dashboard.complication:
        _line(pdf, 6, "No gaps to report for assessable domains.")
    for group in dashboard.complication:
        pdf.set_font("Helvetica", "B", 11)
        _line(
            pdf,
            6,
            f"{group.domain_full_name} ({group.domain_short_code}) - "
            f"{group.met_practices}/{group.total_practices} met",
        )
        pdf.set_font("Helvetica", "I", 9)
        _line(pdf, 5, group.so_what)
        pdf.set_font("Helvetica", "", 9)
        for gap in group.gaps:
            flag = " [AI-proposed, pending review]" if gap.has_pending_ai_proposal else ""
            _line(
                pdf,
                5,
                f"  - {gap.practice_id} ({_mil_label(gap.mil)}): {gap.practice_text}{flag}",
            )
            # ADR-0040: a gap's status/rationale/evidence trail — the
            # actual finding behind it, not just the bare unmet practice —
            # previously computed (ADR-0030) but never rendered anywhere
            # in the export, a real gap found while wiring evidence
            # citation in. finding_rationale is human-authored free text,
            # so it goes through _pdf_safe like every other free-text
            # field this renderer handles.
            if gap.status != PracticeFindingStatus.INSUFFICIENT_EVIDENCE or gap.finding_rationale:
                rationale = (
                    f" - {_pdf_safe(gap.finding_rationale)}" if gap.finding_rationale else ""
                )
                _line(pdf, 5, f"      Status: {_status_label(gap.status)}{rationale}")
            if gap.cited_evidence:
                _line(
                    pdf,
                    5,
                    f"      Evidence: {_evidence_citation_summary(gap.cited_evidence)}",
                )
        pdf.ln(2)
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 13)
    _line(pdf, 8, "Resolution - Prioritized Next Steps")
    pdf.set_font("Helvetica", "", 10)
    if not dashboard.resolution:
        _line(pdf, 6, "No open resolution items.")
    for i, item in enumerate(dashboard.resolution, start=1):
        _line(
            pdf,
            6,
            f"{i}. {item.domain_full_name} ({item.domain_short_code}) - "
            f"{item.missing_count} practice(s) remaining. {item.rationale}",
        )

    return bytes(pdf.output())


_HEADER_FONT = XlsxFont(bold=True)


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def build_xlsx_report(dashboard: DashboardReport) -> bytes:
    """The working-data appendix (a compliance lead's follow-up tool):
    flat, filterable/sortable tables, not prose — deliberately a
    different view of the same DashboardReport than the PDF, not the
    same layout in another format. See ADR-0013.
    """
    wb = Workbook()

    situation_ws = wb.active
    situation_ws.title = "Situation"
    s = dashboard.situation
    generated_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    _write_header(situation_ws, ["Field", "Value"])
    situation_rows = [
        ("Generated", generated_at),
        # Present only once the assessment is finalized; see the PDF
        # header above and Situation.finalization_seal for why it is
        # printed into the export at all.
        *(
            [("Finalization Seal (SHA-256)", s.finalization_seal)]
            if s.finalization_seal
            else []
        ),
        # On every export, not only finalized ones (ADR-0077).
        (f"Report Digest (SHA-256, v{CURRENT_PAYLOAD_VERSION})", report_digest(dashboard)),
        (
            "Check This Report Is Current",
            f"GET /assessments/{s.assessment_id}/report-currency?digest=<the digest above>",
        ),
        ("Assessment Name", s.assessment_name),
        *([("Organization", s.organization_name)] if s.organization_name else []),
        ("Framework", s.framework_name),
        ("Scoring Model", s.scoring_model),
        ("Status", s.status),
        ("Overall Headline", dashboard.overall.headline),
        ("Total Evidence Links", s.total_evidence_links),
        ("Accepted", s.accepted_count),
        ("Edited", s.edited_count),
        ("Rejected", s.rejected_count),
        ("Pending AI Review (not yet scored)", s.pending_ai_review_count),
        ("Unpopulated Domains", ", ".join(s.unpopulated_domains) or "(none)"),
    ]
    # One row per interpretation sentence, rather than all of them joined
    # into a single cell: a spreadsheet reader filters and sorts rows, and
    # a paragraph crammed into one cell is unreadable at the column width
    # any of these sheets use.
    situation_rows += [("What this means", sentence) for sentence in s.so_what]
    for row in situation_rows:
        situation_ws.append(row)
    situation_ws.column_dimensions["A"].width = 32
    situation_ws.column_dimensions["B"].width = 70

    scores_ws = wb.create_sheet("Domain Scores")
    # The header now carries the UNIT. This column held a bare "Score"
    # whose meaning depends on the framework -- an ordinal MIL 0-3 or a
    # 0.0-1.0 fraction -- so a reader sorting it had no way to know which
    # (R-15). Same ambiguity ADR-0066 refused to draw as a bar length,
    # and it was sitting in a spreadsheet the whole time.
    _write_header(scores_ws, ["Domain", _score_column_header(dashboard.overall.scoring_model)])
    for domain, score in dashboard.domain_scores.items():
        scores_ws.append((domain, score))
    scores_ws.column_dimensions["A"].width = 16
    scores_ws.column_dimensions["B"].width = 26

    # Domain completion (ADR-0066/ADR-0069). The XLSX gets the numbers
    # behind the PDF's bars rather than a chart of its own: ADR-0013 made
    # this the flat, sortable working-data appendix on purpose, and warns
    # against pulling the two formats back toward one layout. A reader
    # who wants a chart here has a spreadsheet.
    completion_ws = wb.create_sheet("Domain Completion")
    _write_header(
        completion_ws,
        [
            "Domain Code",
            "Domain Name",
            "Practices Met",
            "Applicable Practices",
            "Completion",
            _score_column_header(dashboard.overall.scoring_model),
            "Blocking MIL",
            "Practices Blocking",
            "What This Means",
        ],
    )
    for entry in dashboard.domain_progress:
        completion_ws.append(
            (
                entry.short_code,
                entry.full_name,
                entry.met_practices,
                entry.total_practices,
                # A real fraction, not a pre-formatted string, so the
                # column sorts and charts as a number in the reader's
                # own spreadsheet.
                entry.met_practices / entry.total_practices,
                entry.score,
                entry.blocking_mil if entry.blocking_mil is not None else "",
                entry.blocking_practice_count if entry.blocking_practice_count is not None else "",
                entry.gate_note or "",
            )
        )
    for cell in completion_ws["E"][1:]:
        cell.number_format = "0%"
    for col, width in zip(
        "ABCDEFGHI", (12, 32, 14, 20, 12, 26, 13, 18, 80), strict=True
    ):
        completion_ws.column_dimensions[col].width = width

    gaps_ws = wb.create_sheet("Gaps")
    _write_header(
        gaps_ws,
        [
            "Domain Code",
            "Domain Name",
            "Practice ID",
            "MIL",
            "Practice Text",
            "AI-Proposed Pending Review",
            "Status",
            "Finding Rationale",
            "Cited Evidence (document id, review status)",
        ],
    )
    for group in dashboard.complication:
        for gap in group.gaps:
            gaps_ws.append(
                (
                    group.domain_short_code,
                    group.domain_full_name,
                    gap.practice_id,
                    _mil_label(gap.mil),
                    gap.practice_text,
                    "Yes" if gap.has_pending_ai_proposal else "No",
                    _status_label(gap.status),
                    gap.finding_rationale or "",
                    _evidence_citation_summary(gap.cited_evidence),
                )
            )
    for col, width in zip(
        "ABCDEFGHI", (12, 32, 14, 8, 80, 24, 20, 60, 50), strict=True
    ):
        gaps_ws.column_dimensions[col].width = width

    resolution_ws = wb.create_sheet("Resolution")
    _write_header(
        resolution_ws, ["Rank", "Domain Code", "Domain Name", "Missing Count", "Rationale"]
    )
    for rank, item in enumerate(dashboard.resolution, start=1):
        resolution_ws.append(
            (
                rank,
                item.domain_short_code,
                item.domain_full_name,
                item.missing_count,
                item.rationale,
            )
        )
    for col, width in zip("ABCDE", (6, 12, 32, 14, 80), strict=True):
        resolution_ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
