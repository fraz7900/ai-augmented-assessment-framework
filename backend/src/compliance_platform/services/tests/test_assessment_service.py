"""Unit tests for the assessment engine, using fakes for both
repositories so the test suite exercises state-machine and
evidence-linking logic without a real SQLite database or LanceDB store.
See services/README.md and tests/README.md.
"""

from __future__ import annotations

from datetime import UTC, datetime

import pytest

from compliance_platform.models.assessment import (
    Assessment,
    AssessmentStatus,
    AssessmentStatusChange,
    Document,
    EvidenceLink,
    EvidenceRequest,
    EvidenceReviewStatus,
    EvidenceSource,
    PracticeFinding,
    PracticeFindingChange,
    PracticeFindingStatus,
    SanitizationApproval,
)
from compliance_platform.models.framework import (
    Domain,
    FrameworkDefinition,
    MilLevelDefinition,
    Objective,
    Practice,
)
from compliance_platform.services.assessment_service import (
    AssessmentFinalizedError,
    AssessmentNotFoundError,
    AssessmentService,
    ChatEngineUnavailableError,
    DocumentNotFoundError,
    EvidenceAlreadyReviewedError,
    EvidenceDocumentNotIngestedError,
    EvidenceLinkNotFoundError,
    EvidenceRequestNotFoundError,
    FrameworkScoringUnavailableError,
    InvalidPracticeReferenceError,
    InvalidReviewDecisionError,
    InvalidStatusTransitionError,
    MappingEngineUnavailableError,
    MissingEvidenceRequestNoteError,
    MissingFindingRationaleError,
    SanitizationApprovalStaleError,
    SanitizationNotApprovedError,
)


class _FakeAssessmentRepository:
    def __init__(self) -> None:
        self._assessments: dict[str, Assessment] = {}
        self._history: dict[str, list[AssessmentStatusChange]] = {}
        self._evidence: dict[str, list[EvidenceLink]] = {}
        self._findings: dict[tuple[str, str], PracticeFinding] = {}
        self._finding_history: dict[str, list[PracticeFindingChange]] = {}
        self._sanitization_approvals: dict[str, list[SanitizationApproval]] = {}
        self._documents: dict[str, Document] = {}
        self._evidence_requests: dict[str, list[EvidenceRequest]] = {}

    def create_assessment(
        self, name: str, framework_name: str, framework_version: str | None = None
    ) -> Assessment:
        assessment = Assessment(
            name=name, framework_name=framework_name, framework_version=framework_version
        )
        self._assessments[assessment.id] = assessment
        self._history[assessment.id] = [
            AssessmentStatusChange(
                assessment_id=assessment.id, from_status=None, to_status=assessment.status
            )
        ]
        self._evidence[assessment.id] = []
        return assessment

    def get_assessment(self, assessment_id: str) -> Assessment | None:
        return self._assessments.get(assessment_id)

    def list_assessments(self) -> list[Assessment]:
        return list(self._assessments.values())

    def update_status(
        self, assessment_id: str, new_status: AssessmentStatus, note: str | None = None
    ) -> Assessment | None:
        assessment = self._assessments.get(assessment_id)
        if assessment is None:
            return None
        previous = assessment.status
        assessment.status = new_status
        self._history[assessment_id].append(
            AssessmentStatusChange(
                assessment_id=assessment_id, from_status=previous, to_status=new_status, note=note
            )
        )
        return assessment

    def status_history(self, assessment_id: str) -> list[AssessmentStatusChange]:
        return list(self._history.get(assessment_id, []))

    def add_evidence_link(self, link: EvidenceLink) -> EvidenceLink:
        self._evidence.setdefault(link.assessment_id, []).append(link)
        return link

    def evidence_for_assessment(self, assessment_id: str) -> list[EvidenceLink]:
        return list(self._evidence.get(assessment_id, []))

    def get_evidence_link(self, evidence_link_id: str) -> EvidenceLink | None:
        for links in self._evidence.values():
            for link in links:
                if link.id == evidence_link_id:
                    return link
        return None

    def update_evidence_link_review(
        self,
        evidence_link_id: str,
        review_status: EvidenceReviewStatus,
        practice_reference: str | None = None,
        note: str | None = None,
    ) -> EvidenceLink | None:
        link = self.get_evidence_link(evidence_link_id)
        if link is None:
            return None
        link.review_status = review_status
        link.reviewed_at = datetime.now(UTC)
        if practice_reference is not None:
            link.practice_reference = practice_reference
        if note is not None:
            link.note = note
        return link

    def set_practice_finding(
        self,
        assessment_id: str,
        practice_reference: str,
        status: PracticeFindingStatus,
        rationale: str,
        set_by: str = "human",
    ) -> PracticeFinding:
        key = (assessment_id, practice_reference)
        existing = self._findings.get(key)
        from_status = existing.status if existing is not None else None
        finding_kwargs = dict(
            assessment_id=assessment_id,
            practice_reference=practice_reference,
            status=status,
            rationale=rationale,
            set_by=set_by,
        )
        if existing is not None:
            finding_kwargs["id"] = existing.id
        finding = PracticeFinding(**finding_kwargs)
        self._findings[key] = finding
        self._finding_history.setdefault(assessment_id, []).append(
            PracticeFindingChange(
                assessment_id=assessment_id,
                practice_reference=practice_reference,
                from_status=from_status,
                to_status=status,
                rationale=rationale,
                set_by=set_by,
            )
        )
        return finding

    def practice_findings_for_assessment(self, assessment_id: str) -> list[PracticeFinding]:
        return [f for (aid, _), f in self._findings.items() if aid == assessment_id]

    def practice_finding_history(
        self, assessment_id: str, practice_reference: str
    ) -> list[PracticeFindingChange]:
        return [
            c
            for c in self._finding_history.get(assessment_id, [])
            if c.practice_reference == practice_reference
        ]

    def create_sanitization_approval(self, approval: SanitizationApproval) -> SanitizationApproval:
        self._sanitization_approvals.setdefault(approval.assessment_id, []).append(approval)
        return approval

    def latest_sanitization_approval(self, assessment_id: str) -> SanitizationApproval | None:
        approvals = self._sanitization_approvals.get(assessment_id, [])
        return approvals[-1] if approvals else None

    def add_document(self, document: Document) -> None:
        # Test-only seeding helper: production Document rows are created
        # via IngestionService, not AssessmentService, so nothing in
        # AssessmentService's own public API creates one to seed with.
        self._documents[document.id] = document

    def get_document(self, document_id: str) -> Document | None:
        return self._documents.get(document_id)

    def document_superseded_by(self, document_id: str) -> Document | None:
        for document in self._documents.values():
            if document.supersedes_document_id == document_id:
                return document
        return None

    def create_evidence_request(self, request: EvidenceRequest) -> EvidenceRequest:
        self._evidence_requests.setdefault(request.assessment_id, []).append(request)
        return request

    def get_evidence_request(self, request_id: str) -> EvidenceRequest | None:
        for requests in self._evidence_requests.values():
            for request in requests:
                if request.id == request_id:
                    return request
        return None

    def evidence_requests_for_assessment(self, assessment_id: str) -> list[EvidenceRequest]:
        return list(self._evidence_requests.get(assessment_id, []))

    def resolve_evidence_request(
        self, request_id: str, resolved_by: str
    ) -> EvidenceRequest | None:
        request = self.get_evidence_request(request_id)
        if request is None:
            return None
        request.resolved_at = datetime.now(UTC)
        request.resolved_by = resolved_by
        return request


class _FakeVectorRepository:
    def __init__(
        self,
        known_documents: dict[str, list[str]] | None = None,
        search_results_by_index: dict[int, list[dict]] | None = None,
        chunk_text: dict[tuple[str, str], str] | None = None,
    ) -> None:
        self._known = known_documents or {}
        self._search_results_by_index = search_results_by_index or {}
        # Chat-specific (Sprint 9): optional real text per (document_id,
        # chunk_id), since chat_service.answer_question reads row["text"]
        # — mapping tests never read this key, so leaving it unset keeps
        # their existing behavior identical.
        self._chunk_text = chunk_text or {}
        self.search_calls: list[tuple[list[float], list[str], int]] = []

    def chunks_for_document(self, document_id: str) -> list[dict]:
        if document_id not in self._known:
            return []
        return [
            {
                "chunk_id": cid,
                "document_id": document_id,
                "text": self._chunk_text.get((document_id, cid), f"text for {cid}"),
            }
            for cid in self._known[document_id]
        ]

    def search_within_documents(
        self, query_vector: list[float], document_ids: list[str], limit: int = 5
    ) -> list[dict]:
        self.search_calls.append((query_vector, document_ids, limit))
        return self._search_results_by_index.get(int(query_vector[0]), [])


class _FakeEmbedder:
    """Returns [index] per input text — see
    services/tests/test_mapping_service.py's identical fake for why."""

    def embed(self, texts: list[str]) -> list[list[float]]:
        return [[float(i)] for i in range(len(texts))]

    @property
    def backend_name(self) -> str:
        return "fake"

    @property
    def dimensions(self) -> int:
        return 1


class _FakeFrameworkRegistry:
    def __init__(self, frameworks: dict[str, FrameworkDefinition] | None = None) -> None:
        self._frameworks = frameworks or {}

    def get(self, name: str) -> FrameworkDefinition | None:
        return self._frameworks.get(name)


def _tiny_framework(name: str = "C2M2") -> FrameworkDefinition:
    """A small, hand-built framework (not real C2M2 content) so
    validation/scoring wiring can be tested without depending on the
    real dataset's specific practice IDs staying stable.
    """
    return FrameworkDefinition(
        name=name,
        full_name="Test Framework",
        version="0",
        source_title="n/a",
        source_publisher="n/a",
        source_date="n/a",
        source_url="n/a",
        retrieved_date="n/a",
        total_practices_in_source=2,
        scoring_model="cumulative_mil",
        mil_levels=[MilLevelDefinition(level=1, name="Initiated", description="n/a")],
        scoring_note="n/a",
        domains=[
            Domain(
                short_code="TEST",
                full_name="Test Domain",
                purpose="n/a",
                practices_populated=True,
                objectives=[
                    Objective(
                        number=1,
                        title="Objective One",
                        practices=[
                            Practice(id="TEST-1a", mil=1, text="practice a"),
                            # An unmet MIL2 practice, so linking only
                            # TEST-1a caps the domain at MIL1 rather than
                            # vacuously satisfying every higher level —
                            # see docs/product/decision_log.md, the
                            # cumulative-scoring correctness rule tested
                            # here is the same one test_scoring_service.py
                            # covers directly.
                            Practice(id="TEST-2a", mil=2, text="practice b, unmet in these tests"),
                        ],
                    )
                ],
            )
        ],
    )


def _make_service(
    known_documents: dict[str, list[str]] | None = None,
    framework_registry: _FakeFrameworkRegistry | None = None,
    embedder: _FakeEmbedder | None = None,
    search_results_by_index: dict[int, list[dict]] | None = None,
    mapping_similarity_threshold: float = 0.5,
    chunk_text: dict[tuple[str, str], str] | None = None,
    chat_similarity_threshold: float = 0.4,
    chat_result_limit: int = 5,
) -> tuple[AssessmentService, _FakeAssessmentRepository, _FakeVectorRepository]:
    assessment_repo = _FakeAssessmentRepository()
    vector_repo = _FakeVectorRepository(known_documents, search_results_by_index, chunk_text)
    service = AssessmentService(
        assessment_repo,
        vector_repo,
        framework_registry=framework_registry,
        embedder=embedder,
        mapping_similarity_threshold=mapping_similarity_threshold,
        chat_similarity_threshold=chat_similarity_threshold,
        chat_result_limit=chat_result_limit,
    )
    return service, assessment_repo, vector_repo


def test_create_assessment_starts_in_draft_with_history_entry() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("Test Assessment", "C2M2")
    assert assessment.status == AssessmentStatus.DRAFT
    history = service.status_history(assessment.id)
    assert len(history) == 1
    assert history[0].from_status is None
    assert history[0].to_status == AssessmentStatus.DRAFT


def test_create_assessment_logs_the_new_assessment_id(caplog: pytest.LogCaptureFixture) -> None:
    # Security hardening (controlled-pilot readiness audit §A.12): "zero
    # logging anywhere in the backend" -- this is the audit trail fix.
    service, _, _ = _make_service()
    with caplog.at_level("INFO", logger="compliance_platform.services.assessment_service"):
        assessment = service.create_assessment("Test Assessment", "C2M2")
    assert any(assessment.id in record.message for record in caplog.records)


def test_create_assessment_pins_framework_version_when_schema_is_loaded() -> None:
    """ADR-0031: the assessment's own record of what version it was
    created against, independent of whatever framework_mapping/*.yaml
    later contains."""
    framework = _tiny_framework()
    assert framework.version == "0"
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("Test Assessment", "C2M2")
    assert assessment.framework_version == "0"


def test_create_assessment_framework_version_is_none_for_an_unrecognized_framework() -> None:
    """No schema loaded for this framework_name at creation time -- the
    same graceful "free text, not yet validated" fallback
    InvalidPracticeReferenceError's docstring documents, not an error."""
    service, _, _ = _make_service()  # no framework_registry configured
    assessment = service.create_assessment("Test Assessment", "Unrecognized Framework")
    assert assessment.framework_version is None


def test_get_assessment_raises_for_unknown_id() -> None:
    service, _, _ = _make_service()
    with pytest.raises(AssessmentNotFoundError):
        service.get_assessment("does-not-exist")


def test_valid_status_transition_updates_status_and_history() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "NIST CSF 2.0")
    updated = service.transition_status(assessment.id, AssessmentStatus.IN_REVIEW, note="ready")
    assert updated.status == AssessmentStatus.IN_REVIEW
    history = service.status_history(assessment.id)
    assert history[-1].to_status == AssessmentStatus.IN_REVIEW
    assert history[-1].note == "ready"


def test_invalid_status_transition_is_rejected() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(InvalidStatusTransitionError):
        service.transition_status(assessment.id, AssessmentStatus.FINALIZED)


def test_finalized_assessment_cannot_transition_further() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    service.transition_status(assessment.id, AssessmentStatus.IN_REVIEW)
    service.transition_status(assessment.id, AssessmentStatus.FINALIZED)
    with pytest.raises(InvalidStatusTransitionError):
        service.transition_status(assessment.id, AssessmentStatus.DRAFT)


def test_link_evidence_succeeds_for_ingested_document() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a", "chunk-b"]})
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(assessment.id, "doc-1", practice_reference="AM-1a")
    assert link.document_id == "doc-1"
    assert link.review_status == EvidenceReviewStatus.ACCEPTED
    assert service.evidence_for_assessment(assessment.id) == [link]


def test_link_evidence_rejects_document_never_ingested() -> None:
    service, _, _ = _make_service(known_documents={})
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(EvidenceDocumentNotIngestedError):
        service.link_evidence(assessment.id, "unknown-doc", practice_reference="AM-1a")


def test_link_evidence_rejects_unknown_chunk_id_for_known_document() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(EvidenceDocumentNotIngestedError):
        service.link_evidence(
            assessment.id, "doc-1", practice_reference="AM-1a", chunk_id="chunk-does-not-exist"
        )


def test_ai_proposed_evidence_defaults_to_pending_review() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(
        assessment.id, "doc-1", practice_reference="AM-1a", source=EvidenceSource.AI_PROPOSED
    )
    assert link.review_status == EvidenceReviewStatus.PENDING


def test_link_evidence_rejected_once_assessment_finalized() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    service.transition_status(assessment.id, AssessmentStatus.IN_REVIEW)
    service.transition_status(assessment.id, AssessmentStatus.FINALIZED)
    with pytest.raises(AssessmentFinalizedError):
        service.link_evidence(assessment.id, "doc-1", practice_reference="AM-1a")


# --- Framework validation and scoring (Sprint 3) ---


def test_link_evidence_with_no_framework_registry_skips_validation() -> None:
    """Backward compatibility with Sprint 2 (Decision D-10): a service
    with no framework_registry configured accepts any practice_reference,
    exactly as it did before Sprint 3.
    """
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(assessment.id, "doc-1", practice_reference="anything-goes")
    assert link.practice_reference == "anything-goes"


def test_link_evidence_accepts_valid_practice_reference_for_known_framework() -> None:
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-a"]}, framework_registry=registry
    )
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(assessment.id, "doc-1", practice_reference="TEST-1a")
    assert link.practice_reference == "TEST-1a"


def test_link_evidence_rejects_unknown_practice_reference_for_known_framework() -> None:
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-a"]}, framework_registry=registry
    )
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(InvalidPracticeReferenceError):
        service.link_evidence(assessment.id, "doc-1", practice_reference="NOT-A-REAL-PRACTICE")


def test_link_evidence_skips_validation_for_framework_with_no_loaded_schema() -> None:
    """An assessment labeled for a framework the registry doesn't have a
    schema for (e.g. "NIST CSF 2.0" before Sprint 4) falls back to
    Sprint 2 free-text behavior rather than blocking evidence linking.
    """
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-a"]}, framework_registry=registry
    )
    assessment = service.create_assessment("A", "NIST CSF 2.0")
    link = service.link_evidence(assessment.id, "doc-1", practice_reference="GV.OC-01")
    assert link.practice_reference == "GV.OC-01"


def test_compute_scores_counts_only_accepted_and_edited_evidence() -> None:
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-a"]}, framework_registry=registry
    )
    assessment = service.create_assessment("A", "C2M2")

    # AI-proposed evidence defaults to PENDING and must not count toward a score.
    service.link_evidence(
        assessment.id, "doc-1", practice_reference="TEST-1a", source=EvidenceSource.AI_PROPOSED
    )
    scores_before_review = service.compute_scores(assessment.id)
    assert scores_before_review["TEST"] == 0

    # Manual evidence defaults to ACCEPTED and must count.
    service.link_evidence(assessment.id, "doc-1", practice_reference="TEST-1a")
    scores_after_manual_link = service.compute_scores(assessment.id)
    assert scores_after_manual_link["TEST"] == 1


def test_compute_scores_raises_when_no_schema_available() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(FrameworkScoringUnavailableError):
        service.compute_scores(assessment.id)


# --- Evidence review workflow (Sprint 5) ---


def test_review_evidence_accept_transitions_pending_to_accepted() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(
        assessment.id, "doc-1", practice_reference="AM-1a", source=EvidenceSource.AI_PROPOSED
    )
    reviewed = service.review_evidence(assessment.id, link.id, EvidenceReviewStatus.ACCEPTED)
    assert reviewed.review_status == EvidenceReviewStatus.ACCEPTED
    assert reviewed.reviewed_at is not None


def test_review_evidence_reject_transitions_pending_to_rejected() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(
        assessment.id, "doc-1", practice_reference="AM-1a", source=EvidenceSource.AI_PROPOSED
    )
    reviewed = service.review_evidence(
        assessment.id, link.id, EvidenceReviewStatus.REJECTED, note="not actually relevant"
    )
    assert reviewed.review_status == EvidenceReviewStatus.REJECTED
    assert reviewed.note == "not actually relevant"


def test_review_evidence_edit_requires_and_validates_corrected_practice_reference() -> None:
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-a"]}, framework_registry=registry
    )
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(
        assessment.id, "doc-1", practice_reference="TEST-1a", source=EvidenceSource.AI_PROPOSED
    )

    with pytest.raises(ValueError):
        service.review_evidence(assessment.id, link.id, EvidenceReviewStatus.EDITED)

    with pytest.raises(InvalidPracticeReferenceError):
        service.review_evidence(
            assessment.id,
            link.id,
            EvidenceReviewStatus.EDITED,
            corrected_practice_reference="NOT-REAL",
        )

    reviewed = service.review_evidence(
        assessment.id,
        link.id,
        EvidenceReviewStatus.EDITED,
        corrected_practice_reference="TEST-2a",
    )
    assert reviewed.review_status == EvidenceReviewStatus.EDITED
    assert reviewed.practice_reference == "TEST-2a"


def test_review_evidence_rejects_reviewing_an_already_reviewed_link() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(assessment.id, "doc-1", practice_reference="AM-1a")  # ACCEPTED
    with pytest.raises(EvidenceAlreadyReviewedError):
        service.review_evidence(assessment.id, link.id, EvidenceReviewStatus.REJECTED)


def test_review_evidence_rejects_unknown_evidence_link() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(EvidenceLinkNotFoundError):
        service.review_evidence(assessment.id, "does-not-exist", EvidenceReviewStatus.ACCEPTED)


def test_review_evidence_rejects_invalid_decision() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(
        assessment.id, "doc-1", practice_reference="AM-1a", source=EvidenceSource.AI_PROPOSED
    )
    with pytest.raises(InvalidReviewDecisionError):
        service.review_evidence(assessment.id, link.id, EvidenceReviewStatus.PENDING)


def test_review_evidence_blocked_on_finalized_assessment() -> None:
    service, _, _ = _make_service(known_documents={"doc-1": ["chunk-a"]})
    assessment = service.create_assessment("A", "C2M2")
    link = service.link_evidence(
        assessment.id, "doc-1", practice_reference="AM-1a", source=EvidenceSource.AI_PROPOSED
    )
    service.transition_status(assessment.id, AssessmentStatus.IN_REVIEW)
    service.transition_status(assessment.id, AssessmentStatus.FINALIZED)
    with pytest.raises(AssessmentFinalizedError):
        service.review_evidence(assessment.id, link.id, EvidenceReviewStatus.ACCEPTED)


# --- AI-proposed mapping engine (Sprint 5) ---


def test_propose_mappings_raises_without_an_embedder_configured() -> None:
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    service, _, _ = _make_service(framework_registry=registry)  # no embedder passed
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(MappingEngineUnavailableError):
        service.propose_mappings(assessment.id)


def test_propose_mappings_raises_without_a_framework_schema() -> None:
    service, _, _ = _make_service(embedder=_FakeEmbedder())
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(FrameworkScoringUnavailableError):
        service.propose_mappings(assessment.id)


def test_propose_mappings_returns_empty_with_no_associated_documents() -> None:
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    service, _, _ = _make_service(framework_registry=registry, embedder=_FakeEmbedder())
    assessment = service.create_assessment("A", "C2M2")
    assert service.propose_mappings(assessment.id) == []


def test_propose_mappings_creates_pending_ai_proposed_links_above_threshold() -> None:
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    # After covering TEST-2a manually, TEST-1a is the sole remaining
    # target practice, so it is index 0 in the batched embed() call.
    search_results = {
        0: [
            {
                "document_id": "doc-1",
                "chunk_id": "chunk-a",
                "_distance": 0.1,
                "text": "matched text",
            }
        ],
    }
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-a"]},
        framework_registry=registry,
        embedder=_FakeEmbedder(),
        search_results_by_index=search_results,
        mapping_similarity_threshold=0.5,
    )
    assessment = service.create_assessment("A", "C2M2")
    # Associate doc-1 with the assessment via a manual link first.
    service.link_evidence(assessment.id, "doc-1", practice_reference="TEST-2a")

    proposed = service.propose_mappings(assessment.id)
    assert len(proposed) == 1
    link = proposed[0]
    assert link.source == EvidenceSource.AI_PROPOSED
    assert link.review_status == EvidenceReviewStatus.PENDING
    assert link.practice_reference == "TEST-1a"
    assert link.confidence is not None
    assert link.confidence > 0.5


def test_propose_mappings_excludes_practices_already_covered() -> None:
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-a"]},
        framework_registry=registry,
        embedder=_FakeEmbedder(),
        search_results_by_index={
            0: [{"document_id": "doc-1", "chunk_id": "chunk-a", "_distance": 0.1, "text": "x"}]
        },
    )
    assessment = service.create_assessment("A", "C2M2")
    service.link_evidence(assessment.id, "doc-1", practice_reference="TEST-1a")
    service.link_evidence(assessment.id, "doc-1", practice_reference="TEST-2a")

    assert service.propose_mappings(assessment.id) == []  # nothing left to propose


def test_propose_mappings_blocked_on_finalized_assessment() -> None:
    registry = _FakeFrameworkRegistry({"C2M2": _tiny_framework()})
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-a"]},
        framework_registry=registry,
        embedder=_FakeEmbedder(),
    )
    assessment = service.create_assessment("A", "C2M2")
    service.link_evidence(assessment.id, "doc-1", practice_reference="TEST-2a")
    service.transition_status(assessment.id, AssessmentStatus.IN_REVIEW)
    service.transition_status(assessment.id, AssessmentStatus.FINALIZED)
    with pytest.raises(AssessmentFinalizedError):
        service.propose_mappings(assessment.id)


# --- Retrieval-only chat (Sprint 8), unit tests added Sprint 9 to close a
# real coverage gap: AssessmentService.answer_question had no direct
# unit test before this, only chat_service.answer_question's own pure-
# function tests and a live API integration test. ---


class _TextKeyedFakeEmbedder:
    """Vectors keyed by exact text match, not call-order index — unlike
    the shared _FakeEmbedder above, which returns [index] per input and
    is unsuitable here: every chat call embeds [question, *chunk_texts]
    together, and [index]-based vectors would make the question (always
    index 0) score an identical, meaningless similarity against every
    chunk regardless of content. Mirrors
    services/tests/test_chat_service.py's fake exactly.
    """

    def __init__(self, vectors_by_text: dict[str, list[float]]) -> None:
        self._vectors_by_text = vectors_by_text

    def embed(self, texts: list[str]) -> list[list[float]]:
        return [self._vectors_by_text.get(text, [0.0, 0.0, 1.0]) for text in texts]

    @property
    def backend_name(self) -> str:
        return "fake"

    @property
    def dimensions(self) -> int:
        return 3


def test_answer_question_raises_for_unknown_assessment() -> None:
    service, _, _ = _make_service(embedder=_FakeEmbedder())
    with pytest.raises(AssessmentNotFoundError):
        service.answer_question("does-not-exist", "any question")


def test_answer_question_raises_without_an_embedder_configured() -> None:
    service, _, _ = _make_service()  # no embedder passed
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(ChatEngineUnavailableError):
        service.answer_question(assessment.id, "any question")


def test_answer_question_returns_empty_with_no_evidence() -> None:
    service, _, _ = _make_service(embedder=_FakeEmbedder())
    assessment = service.create_assessment("A", "C2M2")
    response = service.answer_question(assessment.id, "any question")
    assert response.question == "any question"
    assert response.results == []


def test_answer_question_returns_ranked_hits_from_reviewed_evidence_only() -> None:
    embedder = _TextKeyedFakeEmbedder(
        {
            "which practices cover MFA?": [1.0, 0.0, 0.0],
            "multi-factor authentication is required": [1.0, 0.0, 0.0],
            "unrelated policy text": [0.0, 1.0, 0.0],
        }
    )
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-mfa", "chunk-other"]},
        chunk_text={
            ("doc-1", "chunk-mfa"): "multi-factor authentication is required",
            ("doc-1", "chunk-other"): "unrelated policy text",
        },
        embedder=embedder,
        chat_similarity_threshold=0.5,
    )
    assessment = service.create_assessment("A", "C2M2")
    accepted = service.link_evidence(
        assessment.id, "doc-1", practice_reference="TEST-1a", chunk_id="chunk-mfa"
    )
    # AI-proposed, still pending — must never be answerable until reviewed.
    service.link_evidence(
        assessment.id,
        "doc-1",
        practice_reference="TEST-2a",
        chunk_id="chunk-other",
        source=EvidenceSource.AI_PROPOSED,
    )

    response = service.answer_question(assessment.id, "which practices cover MFA?")
    assert len(response.results) == 1
    result = response.results[0]
    assert result.practice_reference == accepted.practice_reference == "TEST-1a"
    assert result.chunk_id == "chunk-mfa"
    assert result.chunk_text == "multi-factor authentication is required"
    assert result.similarity == 1.0


# --- Practice findings (ADR-0030) ---
#
# The core bug ADR-0030 fixes: before this, "no evidence linked yet for
# TEST-1a" and "evidence was reviewed and shows TEST-1a is NOT actually
# implemented" were indistinguishable — both simply left TEST-1a absent
# from performed_practice_ids, scoring identically. These tests assert
# that gap, directly.


def test_practice_with_zero_evidence_and_practice_with_confirmed_non_compliance_score_identically_without_a_finding() -> (
    None
):
    """Documents the pre-ADR-0030 collapse this feature closes: with NO
    PracticeFinding recorded, a practice with zero evidence and a
    practice whose only evidence was reviewed and REJECTED score
    exactly the same (both simply absent from performed_practice_ids).
    This is the base case the next test's explicit NOT_SATISFIED finding
    is contrasted against — not a bug in this test, but the documented,
    now-escapable default.
    """
    framework = _tiny_framework()
    service, _, _ = _make_service(
        known_documents={"doc-1": ["chunk-1"]},
        framework_registry=_FakeFrameworkRegistry({"C2M2": framework}),
    )
    no_evidence_assessment = service.create_assessment("No evidence", "C2M2")
    rejected_evidence_assessment = service.create_assessment("Rejected evidence", "C2M2")

    link = service.link_evidence(
        rejected_evidence_assessment.id,
        "doc-1",
        practice_reference="TEST-1a",
        source=EvidenceSource.AI_PROPOSED,
    )
    service.review_evidence(
        rejected_evidence_assessment.id,
        link.id,
        decision=EvidenceReviewStatus.REJECTED,
        note="Not actually about this practice.",
    )

    assert service.compute_scores(no_evidence_assessment.id) == service.compute_scores(
        rejected_evidence_assessment.id
    )


def test_practice_finding_not_satisfied_is_distinguishable_from_insufficient_evidence_in_dashboard() -> (
    None
):
    """The actual fix: with an explicit PracticeFinding, the dashboard's
    GapItem.status now distinguishes a practice nobody has looked at
    (INSUFFICIENT_EVIDENCE, the default) from one a reviewer explicitly
    examined and confirmed is not met (NOT_SATISFIED) — even though both
    still correctly count as unmet for MIL/coverage purposes (neither is
    fabricated as "satisfied").
    """
    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("A", "C2M2")

    service.set_practice_finding(
        assessment.id,
        "TEST-1a",
        PracticeFindingStatus.NOT_SATISFIED,
        "Reviewed the submitted policy directly; it does not cover asset tagging at all.",
    )

    dashboard = service.build_dashboard(assessment.id)
    gaps_by_practice = {
        gap.practice_id: gap for group in dashboard.complication for gap in group.gaps
    }
    assert gaps_by_practice["TEST-1a"].status == PracticeFindingStatus.NOT_SATISFIED
    assert "asset tagging" in gaps_by_practice["TEST-1a"].finding_rationale
    # TEST-2a has no finding recorded at all -- still the honest default.
    assert gaps_by_practice["TEST-2a"].status == PracticeFindingStatus.INSUFFICIENT_EVIDENCE
    assert gaps_by_practice["TEST-2a"].finding_rationale is None

    # Neither is fabricated as met: MIL stays 0 (TEST-1a, a MIL1
    # practice, is required for MIL1 and is not satisfied).
    assert service.compute_scores(assessment.id)["TEST"] == 0.0


def test_practice_finding_satisfied_counts_toward_score_without_an_evidence_link() -> None:
    """SATISFIED is authoritative even with zero EvidenceLink rows for
    that practice -- e.g. a documented compensating control a reviewer
    accepts by direct policy review rather than a specific evidence
    upload."""
    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("A", "C2M2")

    service.set_practice_finding(
        assessment.id,
        "TEST-1a",
        PracticeFindingStatus.SATISFIED,
        "Confirmed via direct interview with the control owner; documented in finding notes.",
    )

    assert service.compute_scores(assessment.id)["TEST"] == 1.0  # MIL1 reached


def test_practice_finding_not_applicable_excludes_practice_from_scoring_denominator() -> None:
    """A NOT_APPLICABLE MIL1 practice must not block the domain at MIL0
    forever -- it's removed from the MIL requirement entirely, not
    treated as an unmet blocker."""
    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("A", "C2M2")

    service.set_practice_finding(
        assessment.id,
        "TEST-1a",
        PracticeFindingStatus.NOT_APPLICABLE,
        "This organization has no assets of the type this practice covers.",
    )

    # TEST-1a excluded; TEST-2a (MIL2) is now the only MIL1-tier
    # requirement... but TEST-2a is MIL2, so with TEST-1a excluded there
    # are zero MIL1 practices left to require, and MIL2 remains unmet.
    # Either way, the domain must not be silently blocked *by* the
    # excluded practice.
    assert service.compute_scores(assessment.id)["TEST"] == 1.0

    dashboard = service.build_dashboard(assessment.id)
    gap_practice_ids = {gap.practice_id for group in dashboard.complication for gap in group.gaps}
    assert "TEST-1a" not in gap_practice_ids  # excluded, not a gap
    assert "TEST-2a" in gap_practice_ids  # still genuinely unmet


def test_set_practice_finding_requires_a_rationale() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(MissingFindingRationaleError):
        service.set_practice_finding(
            assessment.id, "TEST-1a", PracticeFindingStatus.NOT_APPLICABLE, ""
        )


def test_set_practice_finding_rejects_unknown_practice_reference() -> None:
    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(InvalidPracticeReferenceError):
        service.set_practice_finding(
            assessment.id, "NOT-A-REAL-PRACTICE", PracticeFindingStatus.SATISFIED, "n/a"
        )


def test_set_practice_finding_blocked_on_finalized_assessment() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    service.transition_status(assessment.id, AssessmentStatus.IN_REVIEW)
    service.transition_status(assessment.id, AssessmentStatus.FINALIZED)
    with pytest.raises(AssessmentFinalizedError):
        service.set_practice_finding(
            assessment.id, "TEST-1a", PracticeFindingStatus.SATISFIED, "n/a"
        )


def test_set_practice_finding_upsert_preserves_history_via_repository() -> None:
    """Confirms the service layer's write actually reaches the
    append-only PracticeFindingChange trail, not just the current-state
    row -- the repository-level mechanics are covered directly in
    repositories/tests/test_assessment_repository.py; this asserts the
    service exposes it correctly end to end.
    """
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")

    service.set_practice_finding(
        assessment.id, "TEST-1a", PracticeFindingStatus.INSUFFICIENT_EVIDENCE, "Nothing yet."
    )
    service.set_practice_finding(
        assessment.id, "TEST-1a", PracticeFindingStatus.NOT_SATISFIED, "Now confirmed absent."
    )

    history = service.practice_finding_history(assessment.id, "TEST-1a")
    assert [h.to_status for h in history] == [
        PracticeFindingStatus.INSUFFICIENT_EVIDENCE,
        PracticeFindingStatus.NOT_SATISFIED,
    ]

    current = service.practice_findings_for_assessment(assessment.id)
    assert len(current) == 1  # upserted, not accumulated as separate rows
    assert current[0].status == PracticeFindingStatus.NOT_SATISFIED


# --- Evidence requests (Sprint 18, ADR-0043) ---


def test_request_more_evidence_requires_a_note() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(MissingEvidenceRequestNoteError):
        service.request_more_evidence(assessment.id, "TEST-1a", "", "priya")


def test_request_more_evidence_rejects_unknown_practice_reference() -> None:
    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(InvalidPracticeReferenceError):
        service.request_more_evidence(
            assessment.id, "NOT-A-REAL-PRACTICE", "need something", "priya"
        )


def test_request_more_evidence_blocked_on_finalized_assessment() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    service.transition_status(assessment.id, AssessmentStatus.IN_REVIEW)
    service.transition_status(assessment.id, AssessmentStatus.FINALIZED)
    with pytest.raises(AssessmentFinalizedError):
        service.request_more_evidence(assessment.id, "TEST-1a", "need something", "priya")


def test_request_more_evidence_creates_an_open_request() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    request = service.request_more_evidence(
        assessment.id, "TEST-1a", "Please provide the current asset inventory.", "priya"
    )
    assert request.practice_reference == "TEST-1a"
    assert request.requested_by == "priya"
    assert request.resolved_at is None
    assert request.resolved_by is None

    listed = service.evidence_requests_for_assessment(assessment.id)
    assert len(listed) == 1
    assert listed[0].id == request.id


def test_multiple_open_requests_can_coexist_for_the_same_practice() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    service.request_more_evidence(assessment.id, "TEST-1a", "first ask", "priya")
    service.request_more_evidence(assessment.id, "TEST-1a", "second ask", "marcus")
    assert len(service.evidence_requests_for_assessment(assessment.id)) == 2


def test_resolve_evidence_request_sets_resolved_fields() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    created = service.request_more_evidence(assessment.id, "TEST-1a", "need X", "priya")

    resolved = service.resolve_evidence_request(assessment.id, created.id, resolved_by="sam")
    assert resolved.resolved_by == "sam"
    assert resolved.resolved_at is not None


def test_resolve_evidence_request_raises_for_unknown_request_id() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(EvidenceRequestNotFoundError):
        service.resolve_evidence_request(assessment.id, "does-not-exist", resolved_by="sam")


def test_resolve_evidence_request_raises_when_request_belongs_to_a_different_assessment() -> None:
    service, _, _ = _make_service()
    a1 = service.create_assessment("A1", "C2M2")
    a2 = service.create_assessment("A2", "C2M2")
    created = service.request_more_evidence(a1.id, "TEST-1a", "need X", "priya")
    with pytest.raises(EvidenceRequestNotFoundError):
        service.resolve_evidence_request(a2.id, created.id, resolved_by="sam")


def test_resolve_evidence_request_blocked_on_finalized_assessment() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    created = service.request_more_evidence(assessment.id, "TEST-1a", "need X", "priya")
    service.transition_status(assessment.id, AssessmentStatus.IN_REVIEW)
    service.transition_status(assessment.id, AssessmentStatus.FINALIZED)
    with pytest.raises(AssessmentFinalizedError):
        service.resolve_evidence_request(assessment.id, created.id, resolved_by="sam")


# --- Sanitization (ADR-0032) ---


def test_preview_sanitization_redacts_pii_in_assessment_name() -> None:
    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("Contact ops@example-utility.com for details", "C2M2")
    preview = service.preview_sanitization(assessment.id)
    assert "ops@example-utility.com" not in preview.sanitized_report.situation.assessment_name


def test_export_sanitized_pdf_blocked_without_any_approval() -> None:
    service, _, _ = _make_service()
    assessment = service.create_assessment("A", "C2M2")
    with pytest.raises(SanitizationNotApprovedError):
        service.generate_dashboard_pdf(assessment.id, sanitized=True)


def test_export_sanitized_pdf_succeeds_after_approval() -> None:
    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("A", "C2M2")
    service.approve_sanitization(assessment.id, custom_terms=[], approved_by="compliance-lead")
    pdf_bytes = service.generate_dashboard_pdf(assessment.id, sanitized=True)
    assert pdf_bytes[:4] == b"%PDF"


def test_export_sanitized_pdf_blocked_after_report_content_changes_post_approval() -> None:
    """The core guarantee: approval is tied to specific content, not a
    standing on/off toggle. Once approved, adding a new finding changes
    the report, so the previously-approved hash no longer matches and
    export must be blocked until re-approved."""
    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("A", "C2M2")
    service.approve_sanitization(assessment.id, custom_terms=[], approved_by="compliance-lead")

    service.set_practice_finding(
        assessment.id, "TEST-1a", PracticeFindingStatus.NOT_SATISFIED, "New finding added."
    )

    with pytest.raises(SanitizationApprovalStaleError):
        service.generate_dashboard_pdf(assessment.id, sanitized=True)

    # Re-approving with the current content succeeds again.
    service.approve_sanitization(assessment.id, custom_terms=[], approved_by="compliance-lead")
    pdf_bytes = service.generate_dashboard_pdf(assessment.id, sanitized=True)
    assert pdf_bytes[:4] == b"%PDF"


def test_approved_custom_terms_are_reused_at_export_time_not_re_supplied() -> None:
    """The approved term list, not whatever the export caller happens to
    pass (export endpoints take no term list at all), governs what a
    sanitized export actually redacts -- so the export always reproduces
    exactly what was reviewed and approved."""
    import io

    from openpyxl import load_workbook

    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("Assessment for Example Utility Co.", "C2M2")
    service.approve_sanitization(
        assessment.id, custom_terms=["Example Utility Co."], approved_by="compliance-lead"
    )
    xlsx_bytes = service.generate_dashboard_xlsx(assessment.id, sanitized=True)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    xlsx_text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Example Utility Co." not in xlsx_text
    assert "ORG-TERM" in xlsx_text


def test_unsanitized_export_is_unaffected_by_sanitization_feature() -> None:
    """sanitized defaults to False -- existing export behavior is
    completely unchanged unless explicitly opted into."""
    framework = _tiny_framework()
    service, _, _ = _make_service(framework_registry=_FakeFrameworkRegistry({"C2M2": framework}))
    assessment = service.create_assessment("Contact ops@example-utility.com", "C2M2")
    pdf_bytes = service.generate_dashboard_pdf(assessment.id)  # sanitized=False, the default
    assert pdf_bytes[:4] == b"%PDF"


# --- Document versioning (Sprint 18, ADR-0039) ---


def test_get_document_detail_raises_for_unknown_document() -> None:
    service, _, _ = _make_service()
    with pytest.raises(DocumentNotFoundError):
        service.get_document_detail("does-not-exist")


def test_get_document_detail_with_no_supersession_relationship() -> None:
    service, assessment_repo, _ = _make_service()
    assessment_repo.add_document(Document(id="doc-1", filename="a.txt", file_type="txt", content_hash="h"))

    detail = service.get_document_detail("doc-1")
    assert detail.id == "doc-1"
    assert detail.supersedes_document_id is None
    assert detail.superseded_by_document_id is None


def test_get_document_detail_surfaces_forward_and_reverse_supersession() -> None:
    service, assessment_repo, _ = _make_service()
    assessment_repo.add_document(Document(id="doc-v1", filename="a.txt", file_type="txt", content_hash="h1"))
    assessment_repo.add_document(
        Document(
            id="doc-v2",
            filename="a.txt",
            file_type="txt",
            content_hash="h2",
            supersedes_document_id="doc-v1",
        )
    )

    v1_detail = service.get_document_detail("doc-v1")
    assert v1_detail.supersedes_document_id is None
    assert v1_detail.superseded_by_document_id == "doc-v2"  # the reviewer-facing signal

    v2_detail = service.get_document_detail("doc-v2")
    assert v2_detail.supersedes_document_id == "doc-v1"
    assert v2_detail.superseded_by_document_id is None
